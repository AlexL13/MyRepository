import xlsxwriter   
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook



def get_info(url):
    wb_data = requests.get(url)
    soup = BeautifulSoup(wb_data.text,'lxml')
    names = soup.select('#contentReviews > h2 > em')
    dates = soup.select('div.meta-value > time')
 #   reviews = soup.select('div.review_quote > div > div.media-body > p')
    infos = soup.select('#movieSynopsis')
    ratings = soup.select('#tomato_meter_link > span.meter-value.superPageFontColor > span')
 #   images = soup.select_one('#poster_link > img')

    d1 = {}
    if names ==[]:
        d1['name'] ='Movie Not Found'
    else:
        for name in names:
            name1 = name.get_text()
            d1['name'] = name1


    if ratings == []:
        d1['rating']='No rating Yet'
    else:
        for rating in ratings:
            rating1 = rating.get_text()
            d1['rating'] = rating1
        
        
    for info in infos:
        info1 = info.get_text()
        d1['info'] = info1


    return d1

    
def cleantext(txt):
    num = txt.find('(')
    text = txt[0:num].strip().replace(':','').replace("'","").replace(' ','_').lower()
    return text

def update_xlsx(dic,dest,r):
    wb = load_workbook(filename = dest,keep_vba=True)   
    ws = wb.get_active_sheet()
    c = 2
    for key in dic:
        ws.cell(row = r,column=c).value = dic[key]
        c = c +1

    wb.save(dest)


def get_url(url):
    L =[]
    wb_data = requests.get(url)
    soup =BeautifulSoup(wb_data.text,'lxml')
    address = soup.find_all('h4')


    for add in address:
        text =add.get_text()
        html = 'https://www.rottentomatoes.com/m/'+cleantext(text)
        L.append(html)
        r = 2
        for link in L:
            update_xlsx(get_info(link),'k360data.xlsm',r)
            r = r + 1
 
    
