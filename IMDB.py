from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook

def cleantext(txt):
    num = txt.find("(")
    text = txt[0:(num-1)]
    return text


def update_xlsx(dic,dest,r):
    wb = load_workbook(filename = dest,keep_vba=True)   
    ws = wb.get_active_sheet()
    c = 2
    for key in dic:
        ws.cell(row = r,column=c).value = dic[key]
        c = c +1

    wb.save(dest)
    
def get_info(url):
    wb_data = requests.get(url)
    soup = BeautifulSoup(wb_data.text,'lxml')
    names = soup.select('#title-overview-widget > div.vital > div.title_block > div > div.titleBar > div.title_wrapper > h1')
    ratings = soup.select('#title-overview-widget > div.vital > div.title_block > div > div.ratings_wrapper > div.imdbRating > div.ratingValue > strong > span')
    links = soup.select('h2 > a')
  
    d1 = {}
    for name in names:
        name1 = cleantext(name.get_text())
        d1['name'] =name1

    for rating in ratings:
        rating1 = rating.get_text()
        d1['rating'] = rating1

    if links ==[]:
        d1['link'] = "Not Avalible"
    else:
        for link in links:
            link1 = link.get('href')
            d1['link'] = 'http://www.imdb.com'+link1

    return d1
    # store data in dictionary {name, rating, links}
 

    
def get_url(url):
    L =[]
    wb_data = requests.get(url)
    soup =BeautifulSoup(wb_data.text,'lxml')
    address = soup.find_all('h4')
    for add in address:
        for urll in add.find_all('a'):
            L.append(urll.get('href'))
        r=2

    for link in L:
        website = 'http://www.imdb.com'+link
        update_xlsx(get_info(website),'k360imdb.xlsm',r)
        r = r + 1
    #get url for all movies
##
##    for add in address:
##        print(add.get_text())
    #get movie names


   
